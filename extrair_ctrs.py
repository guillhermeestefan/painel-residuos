#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extração de CTRs do SP Regula -> dados_ctr.js (para o painel no GitHub Pages).

Roda na nuvem (GitHub Actions), sem depender do seu computador:
  - obras: lidas de obras.csv (Obra, CNPJ, Etapa, Ativo) -- SEM senha (repo público)
  - senha: vem da variável de ambiente SENHA_SPREGULA (GitHub Secret)
  - 1ª execução (sem base ainda): puxa histórico completo desde 01/01/2020
  - execuções seguintes: só o período recente (--dias) e acumula
"""

import argparse, csv, json, os, sys, unicodedata
from datetime import date, timedelta
from pathlib import Path

import smtplib, ssl
from email.mime.text import MIMEText

import openpyxl

BASE = Path(__file__).resolve().parent
OBRAS_CSV = BASE / "obras.csv"
NOVAS_CSV = BASE / "novas_obras.csv"
DOWNLOADS = BASE / "downloads_ctr"
CONSOLIDADO = BASE / "base_consolidada.xlsx"
BASE_JSON = BASE / "base_consolidada.json"
DADOS_JS = BASE / "dados_ctr.js"
ESTADO_EXTRACAO = BASE / "estado_extracao.json"
DIAS_ALERTA = 7   # avisar se uma obra ativa ficar mais de X dias sem extração bem-sucedida
SITE = "https://rcc-spregula.coletas.online/"
SENHA_ENV = os.environ.get("SENHA_SPREGULA", "conx2016")

CLASSE_MAP = {
    "concreto":"A","argamassa":"A","alvenaria":"A","ceramic":"A","cerâmic":"A",
    "solo":"A","terra":"A","escava":"A","entulho":"A","bloco":"A","telha":"A","tijolo":"A",
    "madeira":"B","metal":"B","ferro":"B","aço":"B","papel":"B","papelão":"B",
    "plastic":"B","plástic":"B","gesso":"B","sucata":"B",
    "misturad":"C","lã de vidro":"C","lã de rocha":"C","isopor":"C",
    "tinta":"D","solvente":"D","óleo":"D","oleo":"D","amianto":"D",
    "lâmpada":"D","lampada":"D","bateria":"D","pilha":"D","perigos":"D",
}

def strip_acc(s):
    return "".join(c for c in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(c)).lower()

def classe_de(residuo):
    r = strip_acc(residuo)
    for k, v in CLASSE_MAP.items():
        if strip_acc(k) in r: return v
    return "C"

def etapa_do_residuo(residuo, etapa_cadastro):
    r = strip_acc(residuo)
    if any(k in r for k in ("solo","terrapl","escava")): return "Terraplenagem"
    if "demoli" in r: return "Limpeza do Terreno - Demolição"
    if "stand" in r: return "Stand de Vendas - Demolição"
    return etapa_cadastro or "Não informada"

def normaliza_status(s):
    x = strip_acc(s)
    if "recusad" in x: return "Recusada no Destino Final"
    if "recebida" in x or "baixad" in x: return "Baixada"
    if "transito" in x or "retirada" in x: return "Em Trânsito"
    if "obra" in x: return "Em Obra"
    if "recebimento" in x: return "Pendente Recebimento"
    if "envio" in x: return "Pendente Envio"
    return s or "Baixada"

def _inativa(v):
    return str(v).strip().lower() in ("não","nao","n","false","0")

def _ler_csv(path):
    out = []
    if not path.exists():
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            r = { (k or "").strip().lower(): (v or "").strip() for k, v in r.items() }
            obra = r.get("obra","")
            cnpj = "".join(ch for ch in r.get("login (cnpj)", r.get("cnpj","")) if ch.isdigit())
            senha = r.get("senha","") or SENHA_ENV
            etapa = r.get("etapa atual", r.get("etapa",""))
            ativo = r.get("ativo","Sim")
            if obra and cnpj:
                out.append({"obra":obra,"cnpj":cnpj,"senha":senha,"etapa":etapa,"ativo":ativo})
    return out

PROJECT_ID_FS = "painel-residuos"
def _api_key_fs():
    try:
        txt = open(BASE / "firebase-config.js", encoding="utf-8").read()
    except OSError:
        return ""
    import re as _re
    m = _re.search(r'apiKey:\s*"([^"]+)"', txt)
    return m.group(1) if m else ""
def _fs_decode(v):
    if not isinstance(v, dict): return v
    if 'nullValue' in v: return None
    if 'booleanValue' in v: return v['booleanValue']
    if 'integerValue' in v: return int(v['integerValue'])
    if 'doubleValue' in v: return float(v['doubleValue'])
    if 'stringValue' in v: return v['stringValue']
    if 'mapValue' in v: return {k: _fs_decode(x) for k, x in v['mapValue'].get('fields', {}).items()}
    if 'arrayValue' in v: return [_fs_decode(x) for x in v['arrayValue'].get('values', [])]
    return None
def credenciais_firestore():
    import urllib.request
    key = os.environ.get("FIREBASE_API_KEY", "") or _api_key_fs()
    url = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID_FS}/databases/(default)/documents/painel/credenciais"
    if key:
        url += f"?key={key}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'conx-extrator'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print(f"[aviso] Firestore credenciais indisponivel: {e}")
        return {}
    dados = _fs_decode(payload.get('fields', {}).get('dados'))
    return dados if isinstance(dados, dict) else {}


def _senhas_override():
    """Senhas por obra que diferem da compartilhada, vindas do secret SENHAS_OBRAS
    (JSON {"CNPJ": "senha"}). Mantem senhas fora do repo publico."""
    raw = os.environ.get("SENHAS_OBRAS", "").strip()
    if not raw:
        return {}
    try:
        m = json.loads(raw)
    except Exception as e:
        print(f"[AVISO] SENHAS_OBRAS ignorado (JSON invalido): {e}")
        return {}
    return {"".join(ch for ch in str(k) if ch.isdigit()): str(v) for k, v in m.items()}


def ler_cadastro():
    OVER = _senhas_override()
    FS = credenciais_firestore()
    fs_names = set(FS.keys())
    linhas = []
    for nome, c in FS.items():
        if not isinstance(c, dict):
            continue
        cnpj = "".join(ch for ch in str(c.get("cnpj", "")) if ch.isdigit())
        linhas.append({"obra": nome, "cnpj": cnpj, "senha": str(c.get("senha", "")), "etapa": "", "ativo": str(c.get("ativo", "Sim"))})
    for o in _ler_csv(OBRAS_CSV) + _ler_csv(NOVAS_CSV):
        if o["obra"] in fs_names:
            continue
        linhas.append(o)
    if not linhas:
        sys.exit("[ERRO] Nenhuma fonte de obras (Firestore/CSV).")
    obras, vistos, nomes = [], set(), set()
    for o in linhas:
        nomes.add(o["obra"])
        if _inativa(o["ativo"]) or o["obra"] in vistos:
            continue
        if not o["senha"]:
            continue
        obras.append({"obra": o["obra"], "cnpj": o["cnpj"], "senha": OVER.get(o["cnpj"], o["senha"]), "etapa": o["etapa"]})
        vistos.add(o["obra"])
    if not obras:
        sys.exit("[ERRO] Nenhuma obra ativa.")
    if FS:
        print(f"[INFO] {len(fs_names)} obra(s) do painel (Firestore).")
    return obras, {o["obra"] for o in obras}

def clicar(page, seletores, timeout=8000, obrig=True):
    for sel in seletores:
        try:
            page.locator(sel).first.click(timeout=timeout); return True
        except Exception:
            continue
    if obrig:
        raise RuntimeError(f"Nenhum seletor clicável: {seletores[:2]}")
    return False

def extrair_obra(page, obra, desde, ate, downloads_dir):
    page.goto(SITE, wait_until="domcontentloaded", timeout=60000)
    page.fill("input[type='text']", obra["cnpj"])
    page.fill("input[type='password']", obra["senha"])
    if not clicar(page, ["input[type='submit'][value*='LOGIN' i]",
                         "input[type='button'][value*='LOGIN' i]",
                         "button:has-text('Login')"], obrig=False):
        page.locator("input[type='password']").press("Enter")
    page.wait_for_load_state("networkidle", timeout=60000)

    achou = False
    for nome in ("Gerador","Obra"):
        if clicar(page, [f"text={nome}", f"input[value='{nome}']",
                         f"input[value*='{nome}' i]",
                         f":is(a,button,div,span):has-text('{nome}')"], obrig=False):
            achou = True; break
    if not achou:
        raise RuntimeError("Módulo (Gerador/Obra) não encontrado (login inválido?)")
    page.wait_for_load_state("networkidle", timeout=60000)

    clicar(page, ["text=FECHAR VISUALIZAÇÃO DE AVISOS","input[value*='FECHAR' i]",
                  ":is(a,button):has-text('FECHAR')"], timeout=6000, obrig=False)
    page.wait_for_timeout(600)

    clicar(page, ["a:has-text('Relatórios')","text=Relatórios"], timeout=15000)
    page.wait_for_load_state("networkidle", timeout=60000)
    clicar(page, ["text=registradas",":is(a,div,span):has-text('registradas')"], timeout=15000)
    page.wait_for_timeout(1500)

    d1 = desde.strftime("%d/%m/%Y"); d2 = ate.strftime("%d/%m/%Y")
    page.evaluate(
        """([d1, d2]) => {
            const q = (suf) => document.querySelector("[id$='"+suf+"']");
            const set = (suf, v) => { const el = q(suf); if(el){ el.value = v;
                ['input','change','blur'].forEach(e=>el.dispatchEvent(new Event(e,{bubbles:true}))); } };
            set('ed_DataInicio', d1); set('ed_DataFim', d2);
            const tr = q('ddl_Transportador');
            if (tr) { tr.value = '0'; tr.dispatchEvent(new Event('change',{bubbles:true})); }
        }""", [d1, d2])
    page.wait_for_timeout(500)

    with page.expect_download(timeout=120000) as dl_info:
        if not clicar(page, ["input[id$='bt_Exportar']","input[value*='Exportar' i]"], obrig=False):
            page.locator("text=EXPORTAR").first.click()
    download = dl_info.value
    safe = strip_acc(obra["obra"]).replace(" ","_")[:40] or "obra"
    sufixo = Path(download.suggested_filename).suffix or ".xls"
    dest = downloads_dir / f"ctr_{safe}_{ate:%Y%m%d}{sufixo}"
    download.save_as(str(dest))
    return dest

def _rows_from_df(df):
    return [tuple(df.columns)] + [tuple(r) for r in df.itertuples(index=False, name=None)]

def _ler_html(raw):
    import io, pandas as pd
    try:
        txt = raw.decode("utf-8")
    except UnicodeDecodeError:
        txt = raw.decode("latin-1")
    dfs = pd.read_html(io.StringIO(txt), thousands='.', decimal=',')
    if not dfs:
        raise ValueError("nenhuma tabela HTML encontrada")
    df = max(dfs, key=lambda d: d.shape[0])
    return _rows_from_df(df)

def _ler_spreadsheetml(raw):
    # Excel 2003 XML (SpreadsheetML) - export comum do SP Regula
    from xml.etree import ElementTree as ET
    txt = raw.decode("utf-8", "ignore")
    root = ET.fromstring(txt)
    SS = "{urn:schemas-microsoft-com:office:spreadsheet}"
    rows = []
    for row in root.iter(SS + "Row"):
        cells, col = [], 0
        for cell in row.findall(SS + "Cell"):
            idx = cell.get(SS + "Index")
            if idx:
                col = int(idx) - 1
                while len(cells) < col:
                    cells.append(None)
            data = cell.find(SS + "Data")
            cells.append(data.text if data is not None else None)
            col += 1
        rows.append(tuple(cells))
    if not rows:
        raise ValueError("SpreadsheetML sem linhas")
    return rows

def _ler_xlsx(raw):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    return list(wb.active.iter_rows(values_only=True))

def _ler_xls_bin(raw):
    import io, pandas as pd
    df = pd.read_excel(io.BytesIO(raw), engine="xlrd")
    return _rows_from_df(df)

def _ler_csv_txt(raw):
    import csv as _csv, io as _io
    try:
        txt = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        txt = raw.decode("latin-1")
    amostra = txt[:5000]
    sep = max((";", ","), key=lambda s: amostra.count(s))
    if amostra.count("\t") > amostra.count(sep):
        sep = "\t"
    rows = [tuple(r) for r in _csv.reader(_io.StringIO(txt), delimiter=sep)]
    if len(rows) < 2:
        raise ValueError("CSV com menos de 2 linhas")
    return rows

def carregar_tabela(path):
    """Le o export do SP Regula em qualquer formato entregue: HTML disfarcado
    de .xls, Excel 2003 XML (SpreadsheetML), .xlsx real, .xls binario antigo
    (BIFF) ou CSV. Detecta por conteudo e, se falhar, faz cascata de parsers."""
    path = Path(path)
    raw = open(path, "rb").read()
    if not raw:
        raise ValueError("arquivo vazio (0 bytes) - export nao gerou dados")
    low = raw.lower()
    cabeca = low[:20000]

    ordem = []
    def add(fn):
        if fn not in ordem:
            ordem.append(fn)

    # 1) deteccao por conteudo (nao so pela extensao ou primeiros 4000 bytes)
    if b"mso-application" in cabeca or (b"spreadsheet" in cabeca and b"<workbook" in cabeca):
        add(_ler_spreadsheetml)
    if b"<table" in low or b"<html" in low:
        add(_ler_html)
    if raw[:2] == b"PK":
        add(_ler_xlsx)
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        add(_ler_xls_bin)
    # 2) fallback: cobre qualquer coisa que a deteccao nao pegou
    for fn in (_ler_html, _ler_spreadsheetml, _ler_xlsx, _ler_xls_bin, _ler_csv_txt):
        add(fn)

    erros = []
    for fn in ordem:
        try:
            rows = fn(raw)
            if rows:
                return rows
        except Exception as e:
            erros.append(fn.__name__ + ": " + str(e).splitlines()[0][:60])
    raise ValueError("formato nao reconhecido (magic=" + raw[:48].hex() +
                     ") | tentativas -> " + " ; ".join(erros))

def ler_xls_ctr(path, obra):
    rows = carregar_tabela(path)
    if not rows: return []
    header = [strip_acc(c) for c in rows[0]]
    def col(*names):
        for n in names:
            for i,h in enumerate(header):
                if strip_acc(n)==h: return i
        for n in names:
            for i,h in enumerate(header):
                if strip_acc(n) in h: return i
        return None
    c_num=col("Numero"); c_data=col("DtGuia","data","emiss")
    c_transp=col("tbTransportador_Fantasia","transportador"); c_res=col("TipoResiduo","residuo","material")
    c_qtd=col("VolumeRecebido","Capacidade (M3)","volume","quantidade","m3")
    c_dest=col("Destino","destinac"); c_status=col("Status","situacao")
    c_ender=col("Endereco"); c_pgrcc=col("NumeroPGRCC")
    def cel(r,i): return "" if i is None or i>=len(r) or r[i] is None else str(r[i]).strip()
    regs=[]
    for r in rows[1:]:
        if r is None or all(c is None for c in r): continue
        residuo=cel(r,c_res) or "Não informado"
        try:
            qtd=float(str(r[c_qtd]).replace(",",".")) if c_qtd is not None and r[c_qtd] not in (None,"") else 0.0
        except (ValueError,TypeError): qtd=0.0
        raw=cel(r,c_data); data=raw
        if "/" in raw:
            p=raw.split()[0].split("/")
            if len(p)==3: data=f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
        st=cel(r,c_status)
        regs.append({"obra":obra["obra"],"numero":cel(r,c_num),
            "etapa":etapa_do_residuo(residuo,obra.get("etapa","")),
            "residuo":residuo,"classe":classe_de(residuo),
            "transportador":cel(r,c_transp) or "Não informado",
            "quantidade":round(qtd,2),"destino":cel(r,c_dest) or "Não informado",
            "data":data,"status":normaliza_status(st),
            "divergencia":"divergencia" in strip_acc(st) and "sem" not in strip_acc(st),
            "endereco":cel(r,c_ender),"pgrcc":cel(r,c_pgrcc)})
    return regs

def salvar_consolidado(regs_novos, ativas=None):
    acervo={}
    if BASE_JSON.exists():
        try:
            for r in json.loads(BASE_JSON.read_text(encoding="utf-8")):
                acervo[(r.get("obra"),r.get("numero"))]=r
        except Exception: pass
    for r in regs_novos:
        acervo[(r.get("obra"),r.get("numero"))]=r
    regs=list(acervo.values())
    if ativas is not None:
        regs=[r for r in regs if r.get("obra") in ativas]
    regs=sorted(regs, key=lambda r:(r.get("data",""),r.get("obra","")))
    BASE_JSON.write_text(json.dumps(regs,ensure_ascii=False),encoding="utf-8")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CTRs"
    cols=["numero","data","obra","etapa","residuo","classe","transportador","quantidade","destino","status","divergencia","pgrcc"]
    ws.append([c.capitalize() for c in cols])
    for r in regs: ws.append([r.get(c,"") for c in cols])
    wb.save(CONSOLIDADO)
    meta={"gerado_em":date.today().strftime("%d/%m/%Y"),"total":len(regs),"novos_nesta_rodada":len(regs_novos)}
    DADOS_JS.write_text("window.DADOS_CTR = "+json.dumps(regs,ensure_ascii=False)+";\n"+
                        "window.DADOS_META = "+json.dumps(meta,ensure_ascii=False)+";\n",encoding="utf-8")
    return len(regs)

def _load_estado_extracao():
    if ESTADO_EXTRACAO.exists():
        try: return json.loads(ESTADO_EXTRACAO.read_text(encoding="utf-8"))
        except Exception: return {}
    return {}

def _enviar_alerta(assunto, corpo_html):
    """Envia alerta por e-mail usando os mesmos secrets MAIL_* do fluxo. Sem SMTP configurado, só imprime."""
    server=os.environ.get("MAIL_SERVER",""); user=os.environ.get("MAIL_USERNAME","")
    pwd=os.environ.get("MAIL_PASSWORD",""); mail_from=os.environ.get("MAIL_FROM",user)
    to=[e.strip() for e in (os.environ.get("MAIL_TO","") or "").replace(";",",").split(",") if e.strip()]
    if not (server and user and pwd and to):
        print(f"   [PRÉVIA — e-mail não enviado, SMTP/MAIL_TO ausente] {assunto} -> {to}"); return
    try: port=int((os.environ.get("MAIL_PORT") or "587").strip())
    except Exception: port=587
    msg=MIMEText(corpo_html,"html","utf-8"); msg["Subject"]=assunto
    msg["From"]=mail_from; msg["To"]=", ".join(to)
    ctx=ssl.create_default_context()
    try:
        if port==465:
            with smtplib.SMTP_SSL(server,port,context=ctx,timeout=30) as sv:
                sv.login(user,pwd); sv.sendmail(mail_from,to,msg.as_string())
        else:
            with smtplib.SMTP(server,port,timeout=30) as sv:
                sv.ehlo(); sv.starttls(context=ctx); sv.login(user,pwd); sv.sendmail(mail_from,to,msg.as_string())
        print(f"   [ALERTA ENVIADO] {assunto} -> {', '.join(to)}")
    except Exception as e:
        print(f"   [ALERTA FALHOU AO ENVIAR] {e}")

def atualizar_estado_e_alertar(resultados, ativas, ate):
    """Registra o último sucesso por obra e avisa quem passou de DIAS_ALERTA dias sem extrair.
    Os dados anteriores NUNCA são apagados por falha: salvar_consolidado só acrescenta/atualiza."""
    hoje = ate.strftime("%Y-%m-%d")
    estado = _load_estado_extracao()
    for obra, res in resultados.items():
        e = estado.get(obra, {})
        if res["ok"]:
            e["ultimo_sucesso"]=hoje; e["falhas_seguidas"]=0; e["ultimo_erro"]=""
        else:
            e["falhas_seguidas"]=int(e.get("falhas_seguidas",0))+1
            e["ultimo_erro"]=res.get("erro","")
            e.setdefault("ultimo_sucesso", e.get("ultimo_sucesso",""))
        estado[obra]=e
    ESTADO_EXTRACAO.write_text(json.dumps(estado,ensure_ascii=False,indent=1),encoding="utf-8")

    atrasadas=[]
    for obra in ativas:
        e=estado.get(obra,{}); us=e.get("ultimo_sucesso","")
        if not us:
            atrasadas.append((obra,"nunca extraída com sucesso",e.get("ultimo_erro","")))
        else:
            try: dias=(ate-date.fromisoformat(us)).days
            except Exception: dias=0
            if dias>DIAS_ALERTA:
                atrasadas.append((obra,f"{dias} dias sem sucesso (último em {us})",e.get("ultimo_erro","")))
    if atrasadas:
        linhas="".join(f"<li><b>{o}</b> — {q}{(' · '+err) if err else ''}</li>" for o,q,err in atrasadas)
        html=(f"<p>As obras abaixo estão há mais de {DIAS_ALERTA} dias sem extração de CTRs "
              f"bem-sucedida no SP Regula:</p><ul>{linhas}</ul>"
              f"<p style='color:#666'>Os dados anteriores dessas obras foram <b>preservados</b> no painel — "
              f"apenas não houve atualização. Verifique login/senha ou disponibilidade do SP Regula.</p>")
        _enviar_alerta(f"[Painel de Resíduos] {len(atrasadas)} obra(s) sem extração há +{DIAS_ALERTA} dias", html)
        print(f"[ALERTA] {len(atrasadas)} obra(s) atrasada(s): "+", ".join(o for o,_,_ in atrasadas))
    else:
        print(f"[OK] Nenhuma obra ativa com mais de {DIAS_ALERTA} dias sem extração.")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--dias", type=int, default=30)
    ap.add_argument("--desde", help="dd/mm/aaaa")
    ap.add_argument("--headful", action="store_true")
    args=ap.parse_args()

    from playwright.sync_api import sync_playwright
    DOWNLOADS.mkdir(exist_ok=True)
    ate=date.today()
    # 1ª execução (sem base): histórico completo. Depois: janela recente.
    if args.desde:
        d,m,a=args.desde.split("/"); desde=date(int(a),int(m),int(d))
    elif not BASE_JSON.exists():
        desde=date(2020,1,1); print("[INFO] 1ª execução: puxando histórico completo desde 01/01/2020")
    else:
        desde=ate-timedelta(days=args.dias)

    obras, ativas = ler_cadastro()
    print(f"[INFO] {len(obras)} obra(s) ativa(s) | período {desde:%d/%m/%Y} a {ate:%d/%m/%Y}")

    todos=[]
    resultados={}
    with sync_playwright() as p:
        browser=p.chromium.launch(headless=not args.headful)
        import time
        for o in obras:
            ok = False
            ultimo = ""
            for tent in range(1, 4):
                ctx = browser.new_context(accept_downloads=True); page = ctx.new_page()
                try:
                    print(f"  -> {o['obra']} (CNPJ {o['cnpj']}) tent {tent}/3 ...", end=" ", flush=True)
                    arq = extrair_obra(page, o, desde, ate, DOWNLOADS)
                    regs = ler_xls_ctr(arq, o); todos.extend(regs)
                    print(f"OK ({len(regs)} CTRs)")
                    ok = True
                except Exception as e:
                    ultimo = str(e).splitlines()[0][:90]
                    print(f"tent {tent} falhou: {ultimo}")
                finally:
                    ctx.close()
                if ok:
                    break
                time.sleep(6)
            if not ok:
                print(f"  [FALHOU] {o['obra']} apos 3 tentativas: {ultimo}")
            resultados[o['obra']] = {"ok": ok, "erro": ultimo}
            time.sleep(2)
        browser.close()

    # Preserva os dados anteriores: salvar_consolidado só acrescenta/atualiza, nunca zera obras que falharam.
    total=salvar_consolidado(todos, ativas)
    print(f"[OK] +{len(todos)} CTRs nesta rodada | base agora com {total} CTRs")
    # Registra saúde da extração por obra e avisa se alguma passar de DIAS_ALERTA dias sem sucesso.
    try:
        atualizar_estado_e_alertar(resultados, ativas, ate)
    except Exception as e:
        print(f"[AVISO] não foi possível atualizar estado/alerta de extração: {e}")

if __name__ == "__main__":
    main()
